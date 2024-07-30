import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font
from openpyxl import load_workbook
import threading

searching = False  # Global variable to control the search

def search_id_in_excel(directory, id_to_find, fuzzy_match):
    results = []
    unreadable_files = []
    for root, dirs, files in os.walk(directory):
        if not searching:  # Check if search has been stopped
            break
        for file in files:
            if not searching:  # Check if search has been stopped
                break
            if file.startswith("~$"):
                continue  # Skip temporary Excel files
            if file.endswith(".xlsx"):  # Adjust file extension as needed
                file_path = os.path.join(root, file)
                try:
                    # Open workbook to read formulas
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=False)
                    formula_data = {}
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value is not None and hasattr(cell, 'coordinate'):
                                    cell_formula = str(cell.value) if cell.data_type == 'f' else ""
                                    formula_data[cell.coordinate] = cell_formula
                    workbook.close()

                    # Open workbook to read values
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value is not None and hasattr(cell, 'coordinate'):
                                    cell_value = str(cell.value)
                                    cell_formula = formula_data.get(cell.coordinate, "")
                                    if fuzzy_match:
                                        if id_to_find.lower() in cell_value.lower() or id_to_find.lower() in cell_formula.lower():
                                            rel_path = os.path.relpath(file_path, directory).replace('/', '\\')
                                            results.append((rel_path, sheet_name, cell.coordinate, cell_value, cell_formula))
                                    else:
                                        if cell_value == id_to_find or cell_formula == id_to_find:
                                            rel_path = os.path.relpath(file_path, directory).replace('/', '\\')
                                            results.append((rel_path, sheet_name, cell.coordinate, cell_value, cell_formula))
                    workbook.close()
                except Exception as e:
                    print(f"Error processing {file}: {e}")
                    unreadable_files.append(file_path.replace('/', '\\'))
                    update_unreadable_files(file_path.replace('/', '\\'))
                    continue  # Skip to the next file if there's an error
    return results, unreadable_files

def update_unreadable_files(file_path):
    global error_label_added
    if not error_label_added:
        error_label = ttk.Label(result_frame, text="以下文件无法读取:", foreground='red')
        error_label.pack(pady=5)
        error_label_added = True
    
    error_file_label = ttk.Label(result_frame, text=file_path, foreground='red')
    error_file_label.pack(pady=2)
    root.update_idletasks()

def select_directory():
    directory_path = filedialog.askdirectory()
    if directory_path:
        directory_path_entry.delete(0, tk.END)
        directory_path_entry.insert(0, directory_path.replace('/', '\\'))

def search_button_click():
    global searching
    global error_label_added
    error_label_added = False
    if search_button['text'] == "停止搜索":
        searching = False
        search_button['text'] = "搜索 (Enter)"
        directory_path_entry.config(state='normal')
        id_to_search_entry.config(state='normal')
        select_directory_button.config(state='normal')
        return

    directory_path = directory_path_entry.get()
    id_to_search = id_to_search_entry.get()
    fuzzy_match = fuzzy_match_var.get() == "模糊搜索"
    
    if not directory_path or not id_to_search:
        messagebox.showerror("错误", "请输入目录路径和要搜索的ID。")
        return

    clear_previous_results()

    searching = True
    search_button['text'] = "停止搜索"
    directory_path_entry.config(state='disabled')
    id_to_search_entry.config(state='disabled')
    select_directory_button.config(state='disabled')
    
    wait_label = tk.Label(root, text="搜索中，请稍等...")
    wait_label.pack(pady=10)
    root.update_idletasks()
    
    search_thread = threading.Thread(target=perform_search, args=(directory_path, id_to_search, fuzzy_match, wait_label))
    search_thread.start()

def perform_search(directory_path, id_to_search, fuzzy_match, wait_label):
    try:
        search_results, unreadable_files = search_id_in_excel(directory_path, id_to_search, fuzzy_match)
        wait_label.pack_forget()
        search_button['text'] = "搜索 (Enter)"
        directory_path_entry.config(state='normal')
        id_to_search_entry.config(state='normal')
        display_results(search_results, unreadable_files, fuzzy_match)
    except Exception as e:
        wait_label.pack_forget()
        search_button['text'] = "搜索 (Enter)"
        directory_path_entry.config(state='normal')
        id_to_search_entry.config(state='normal')
        messagebox.showerror("错误", f"发生错误: {e}")

def display_results(results, unreadable_files, fuzzy_match):
    for widget in result_frame.winfo_children():
        widget.destroy()
    
    if unreadable_files:
        error_label = ttk.Label(result_frame, text="以下文件无法读取:", foreground='red')
        error_label.pack(pady=5)
        for file in unreadable_files:
            error_file_label = ttk.Label(result_frame, text=file, foreground='red')
            error_file_label.pack(pady=2)

    if not results:
        result_label = ttk.Label(result_frame, text="未找到结果。")
        result_label.pack(pady=10)
    else:
        columns = ("文件路径", "工作簿名称", "位置", "值", "公式")
        
        tree_frame = ttk.Frame(result_frame)
        tree_frame.pack(fill="both", expand=True)
        tree_frame.pack_propagate(False)  # Ensure scrollbar always shows

        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        tree.heading("文件路径", text="文件路径")
        tree.heading("工作簿名称", text="工作簿名称")
        tree.heading("位置", text="位置")
        tree.heading("值", text="值")
        tree.heading("公式", text="公式")
        
        # Center align all cell contents
        tree.tag_configure('center', anchor='center')
        
        # Insert results into the treeview
        for result in results:
            tree.insert("", "end", values=result, tags='center')
        
        tree.pack(side="left", fill="both", expand=True)
        
        # Add vertical scrollbar to the treeview
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Remove horizontal scrollbar
        tree.configure(xscroll=ttk.Scrollbar(tree_frame, orient="horizontal", command=lambda *args: None).set)
        
        # Enable mouse wheel scrolling
        def mouse_wheel(event):
            tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        tree.bind_all("<MouseWheel>", mouse_wheel)
        
        # Double click event to open Excel file
        def on_double_click(event):
            item = tree.selection()[0]
            file_path = tree.item(item, 'values')[0]
            abs_path = os.path.normpath(os.path.join(directory_path_entry.get(), file_path))
            
            try:
                # Disable the search button and main window to prevent re-triggering the search
                search_button.config(state='disabled')
                root.config(cursor="wait")
                root.update_idletasks()
                
                os.startfile(abs_path)
                messagebox.showinfo("打开Excel", f"正在打开 {abs_path}...")
            except Exception as e:
                messagebox.showerror("错误", f"打开 {abs_path} 时发生错误: {e}")
            finally:
                # Re-enable the search button and main window
                search_button.config(state='normal')
                root.config(cursor="")
                root.update_idletasks()
        
        tree.bind("<Double-1>", on_double_click)
        tree.bind("<Return>", on_double_click)
        
        # Adjust column widths
        max_width = 300  # Set a maximum width for columns
        total_width = 0
        for col in columns:
            max_col_width = max([font.Font().measure(str(tree.set(child, col))) for child in tree.get_children()])
            col_width = min(max_col_width + 20, max_width)  # Add some padding
            tree.column(col, width=col_width, stretch=False)
            total_width += col_width

        # Adjust the Treeview width to fit the window width minus scrollbar width
        tree_frame.update_idletasks()
        window_width = root.winfo_width()
        scrollbar_width = scrollbar.winfo_width()
        tree_width = min(total_width, window_width - scrollbar_width)
        tree_frame.config(width=tree_width)

def clear_previous_results():
    for widget in result_frame.winfo_children():
        widget.destroy()

# Function to get DPI and adjust font size
def get_dpi_and_adjust_font():
    # Get the screen's DPI
    dpi = root.winfo_fpixels('1i')
    # Adjust font size based on DPI
    base_font_size = 12
    adjusted_font_size = int(base_font_size * (dpi / 96))  # Assuming 96 DPI as the base
    return adjusted_font_size

# Create main window
root = tk.Tk()
root.title("多Excel表查找 - By Cheuksing (v1.0)")

# Set application icon
try:
    root.iconbitmap('E:\测试文件夹\logo192.ico')
except:
    pass

# Enable high DPI awareness (Windows only)
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass
# Adjust font size based on DPI
adjusted_font_size = get_dpi_and_adjust_font()
default_font = font.Font(family="Microsoft YaHei", size=adjusted_font_size)
root.option_add("*Font", default_font)
# 配置 ttk.Style
style = ttk.Style()
style.configure('TButton', font=(default_font.actual()['family'], adjusted_font_size))
style.configure('TRadiobutton', font=(default_font.actual()['family'], adjusted_font_size))
style.configure('Treeview', font=(default_font.actual()['family'], adjusted_font_size))
style.configure('Treeview.Heading', font=(default_font.actual()['family'], adjusted_font_size))

# Function to center the window on the screen
def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

# Adjust window size and position
window_width = 800
window_height = 600
center_window(root, window_width, window_height)

# Create directory path entry
directory_path_label = ttk.Label(root, text="目录路径:")
directory_path_label.pack(pady=10, anchor="w")

directory_path_frame = ttk.Frame(root)
directory_path_frame.pack(fill="x", padx=10, pady=(0, 10))

directory_path_entry = ttk.Entry(directory_path_frame, width=50)
directory_path_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))

select_directory_button = ttk.Button(directory_path_frame, text="选择文件夹", command=select_directory)
select_directory_button.pack(side=tk.LEFT)

# Create ID to search entry
id_to_search_label = ttk.Label(root, text="搜索单元格内容（支持显示值和公式）：")
id_to_search_label.pack(pady=10, anchor="w")

id_to_search_entry = ttk.Entry(root, width=50)
id_to_search_entry.pack(fill="x", padx=10)

# Create match type selection
match_type_frame = ttk.Frame(root)
match_type_frame.pack(fill="x", padx=10, pady=10)

fuzzy_match_var = tk.StringVar(value="精确匹配")

exact_match_rb = ttk.Radiobutton(match_type_frame, text="精确匹配 (完全匹配，区分大小写)", variable=fuzzy_match_var, value="精确匹配", style="TRadiobutton")
exact_match_rb.pack(fill="x",padx=10)

fuzzy_match_rb = ttk.Radiobutton(match_type_frame, text="模糊匹配 (部分匹配，不区分大小写)", variable=fuzzy_match_var, value="模糊搜索", style="TRadiobutton")
fuzzy_match_rb.pack(fill="x",padx=10)

# Create Search button
search_button = ttk.Button(root, text="搜索 (Enter)", command=search_button_click)
search_button.pack(pady=20)

# 绑定回车键事件到整个窗口
root.bind("<Return>", lambda event: search_button_click())

# Create frame for displaying results
result_frame = ttk.Frame(root)
result_frame.pack(fill="both", expand=True, padx=10, pady=10)

# Start the main loop
root.mainloop()
#pyinstaller --onefile --windowed --icon="E:\测试文件夹\logo192.ico" "E:\测试文件夹\test4_ui.py" -F